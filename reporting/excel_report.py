import os
import platform
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# ---------------------------------------------------------
# FAILURE CLASSIFICATION (IMPROVED)
# ---------------------------------------------------------
def classify_failure(error_text):
    if not error_text:
        return ""

    error_text = str(error_text)

    if "AssertionError" in error_text:
        return "ASSERTION"

    if "409" in error_text:
        return "BUSINESS_RULE"

    if "CHECK constraint" in error_text or "IntegrityError" in error_text:
        return "DATA_INTEGRITY"

    if "pyodbc" in error_text or "SQL" in error_text:
        return "DB_ERROR"

    if "Timeout" in error_text:
        return "TIMEOUT"

    if "HTTP" in error_text or "status_code" in error_text:
        return "API_ERROR"

    return "SYSTEM_ERROR"


# ---------------------------------------------------------
# MAIN REPORT GENERATOR
# ---------------------------------------------------------
def generate_report(session):
    os.makedirs("reports", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    headers = [
        "Test Name",
        "Markers",
        "Outcome",
        "Duration (ms)",
        "Failure Type",
        "Error Message",
        "Screenshot",
        "Environment",
        "Build ID",
        "Timestamp"
    ]

    ws.append(headers)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    pass_count = 0
    fail_count = 0
    skip_count = 0

    build_id = os.getenv("BUILD_ID", "LOCAL_RUN")
    environment = os.getenv("ENV_NAME", platform.system())

    for item in session.items:

        outcome = "UNKNOWN"
        duration = ""
        error_msg = ""
        failure_type = ""
        screenshot = getattr(item, "screenshot_path", "")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        markers = ",".join([m.name for m in item.iter_markers()])

        rep = getattr(item, "rep_call", None)

        if rep:
            duration = round(rep.duration * 1000, 2)

            if rep.passed:
                outcome = "PASSED"
                pass_count += 1

            elif rep.failed:
                outcome = "FAILED"
                fail_count += 1

                # Clean error message (avoid massive traceback noise)
                error_msg = str(rep.longrepr)
                error_msg = error_msg.split("E       ")[-1] if "E       " in error_msg else error_msg[:500]

                failure_type = classify_failure(error_msg)

            elif rep.skipped:
                outcome = "SKIPPED"
                skip_count += 1

        row = [
            item.name,
            markers,
            outcome,
            duration,
            failure_type,
            error_msg,
            screenshot,
            environment,
            build_id,
            timestamp
        ]

        ws.append(row)

        # Highlight failures
        if outcome == "FAILED":
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = red_fill

    # ---------------------------------------------------------
    # SUMMARY SHEET
    # ---------------------------------------------------------
    summary = wb.create_sheet("Summary")

    summary.append(["Total Tests", len(session.items)])
    summary.append(["Passed", pass_count])
    summary.append(["Failed", fail_count])
    summary.append(["Skipped", skip_count])
    summary.append(["Build ID", build_id])
    summary.append(["Environment", environment])
    summary.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    timestamp_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = f"reports/test_report_{timestamp_file}.xlsx"

    wb.save(report_path)

    print(f"\nAdvanced Excel Report Generated: {report_path}")